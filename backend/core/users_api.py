"""REST API Endpoints for Auth, Users Management, RBAC Roles, and Audit Logs."""
from __future__ import annotations

import csv
import datetime
import io
import time
import uuid
from typing import Any, Dict, List, Optional
from fastapi import APIRouter, Depends, Request, Response, status
from pydantic import BaseModel, EmailStr

from backend.core.exceptions import NMSError, NotFoundError, ValidationError, AuthenticationError, PermissionDeniedError
from backend.core.auth import (
    CurrentUser,
    clear_permissions_cache,
    create_access_token,
    decode_access_token,
    get_current_user,
    is_ip_whitelisted,
    require_permission,
)
from backend.core.audit import log_audit_event
from backend.core.database import get_db_connection, hash_password, verify_password
from backend.core.i18n import get_lang, tr
from backend.core.plugin.registry import get_security_settings, save_security_settings
from backend.core.mfa import (
    generate_totp_secret,
    get_totp_uri,
    generate_qr_svg,
    verify_totp_code,
)

router = APIRouter(prefix="/api", tags=["auth_users_rbac"])

# Хранилище билетов для второго шага MFA (ticket -> {user_id, username, expires_at})
mfa_tickets: Dict[str, Dict[str, Any]] = {}


# ── Схемы данных (Pydantic) ──────────────────────────────────────────
class LoginRequest(BaseModel):
    username: str
    password: str


class LoginResponse(BaseModel):
    token: str = ""
    user: Dict[str, Any] = {}
    must_change_password: bool = False
    mfa_required: bool = False
    mfa_ticket: Optional[str] = None
    mfa_setup_required: bool = False
    qr_code: Optional[str] = None
    secret: Optional[str] = None


class MfaVerifyRequest(BaseModel):
    mfa_ticket: str
    code: str


class MfaEnableRequest(BaseModel):
    secret: str
    code: str


class UserCreateRequest(BaseModel):
    username: str
    password: str
    full_name: str
    email: Optional[str] = None
    title: Optional[str] = None
    uid: Optional[str] = None
    role_id: str
    is_active: bool = True
    must_change_password: Optional[bool] = None


class UserUpdateRequest(BaseModel):
    full_name: Optional[str] = None
    email: Optional[str] = None
    title: Optional[str] = None
    role_id: Optional[str] = None
    is_active: Optional[bool] = None
    password: Optional[str] = None
    must_change_password: Optional[bool] = None


class SelfUpdateRequest(BaseModel):
    full_name: Optional[str] = None
    email: Optional[str] = None
    avatar: Optional[str] = None


class PasswordChangeRequest(BaseModel):
    old_password: str
    new_password: str


class RoleCreateUpdateRequest(BaseModel):
    name: str
    description: Optional[str] = ""
    permission_ids: List[str] = []


# ── 1. Auth Endpoints ────────────────────────────────────────────────
@router.post("/auth/login", response_model=LoginResponse)
async def login(body: LoginRequest, request: Request):
    """Вход пользователя в систему."""
    sec_settings = get_security_settings()

    ip_whitelist = sec_settings.get("ip_whitelist", "")
    if ip_whitelist and request and request.client:
        client_ip = request.client.host
        if not is_ip_whitelisted(client_ip, ip_whitelist):
            raise PermissionDeniedError(message=tr(request, "ip_access_denied", client_ip=client_ip), code="IP_ACCESS_DENIED")

    conn = get_db_connection()
    try:
        user = conn.execute(
            """
            SELECT u.id, u.username, u.full_name, u.email, u.uid, u.avatar, u.hashed_password, u.is_active, u.role_id, r.name as role_name,
                   u.must_change_password, u.failed_login_attempts, u.locked_until, u.mfa_enabled, u.mfa_secret
            FROM users u
            JOIN roles r ON u.role_id = r.id
            WHERE u.username = ?
            """,
            (body.username,),
        ).fetchone()

        now = datetime.datetime.now(datetime.timezone.utc)
        if user and user["locked_until"]:
            try:
                locked_until_dt = datetime.datetime.fromisoformat(str(user["locked_until"]))
                if locked_until_dt.tzinfo is None:
                    locked_until_dt = locked_until_dt.replace(tzinfo=datetime.timezone.utc)
                if now < locked_until_dt:
                    log_audit_event(
                        user_id=user["id"],
                        username=body.username,
                        action="auth.login_failed",
                        resource="auth",
                        details=tr(request, "login_attempt_locked"),
                        ip_address=request.client.host if request.client else None,
                    )
                    raise NMSError(message=tr(request, "account_temporarily_locked"), status_code=429, code="ACCOUNT_TEMPORARILY_LOCKED")
            except NMSError:
                raise
            except Exception:
                pass

        if not user or not verify_password(body.password, user["hashed_password"]):
            if user:
                failed_cnt = (user["failed_login_attempts"] or 0) + 1
                max_attempts = int(sec_settings.get("max_login_attempts", 5))
                lockout_duration = int(sec_settings.get("lockout_duration", 30))
                if failed_cnt >= max_attempts:
                    locked_until_time = (now + datetime.timedelta(minutes=lockout_duration)).isoformat()
                    conn.execute(
                        "UPDATE users SET failed_login_attempts = ?, locked_until = ? WHERE id = ?",
                        (failed_cnt, locked_until_time, user["id"]),
                    )
                    conn.commit()
                    log_audit_event(
                        user_id=user["id"],
                        username=body.username,
                        action="auth.login_lockout",
                        resource="auth",
                        details=tr(request, "account_locked_duration_audit", lockout_duration=lockout_duration),
                        ip_address=request.client.host if request.client else None,
                    )
                    raise NMSError(message=tr(request, "account_locked_duration_detail", lockout_duration=lockout_duration), status_code=429, code="ACCOUNT_LOCKED_DURATION")
                else:
                    conn.execute("UPDATE users SET failed_login_attempts = ? WHERE id = ?", (failed_cnt, user["id"]))
                    conn.commit()

            log_audit_event(
                user_id=None,
                username=body.username,
                action="auth.login_failed",
                resource="auth",
                details=tr(request, "invalid_credentials"),
                ip_address=request.client.host if request.client else None,
            )
            raise AuthenticationError(message=tr(request, "invalid_credentials"), code="INVALID_CREDENTIALS")

        if not user["is_active"]:
            raise PermissionDeniedError(message=tr(request, "account_locked"), code="ACCOUNT_LOCKED")

        # Проверка MFA
        force_mfa = bool(sec_settings.get("force_mfa", False))
        user_mfa_enabled = bool(user["mfa_enabled"] if "mfa_enabled" in user.keys() and user["mfa_enabled"] else False)
        user_mfa_secret = user["mfa_secret"] if "mfa_secret" in user.keys() else None

        if user_mfa_enabled and user_mfa_secret:
            mfa_ticket = f"mfat_{uuid.uuid4().hex}"
            mfa_tickets[mfa_ticket] = {
                "user_id": user["id"],
                "username": user["username"],
                "mfa_secret": user_mfa_secret,
                "is_setup": False,
                "expires_at": time.time() + 300,
            }
            return {
                "token": "",
                "mfa_required": True,
                "mfa_setup_required": False,
                "mfa_ticket": mfa_ticket,
                "must_change_password": bool(user["must_change_password"]),
                "user": {},
            }
        elif force_mfa:
            setup_secret = generate_totp_secret()
            totp_uri = get_totp_uri(setup_secret, user["username"], issuer="NMS WebUI")
            qr_svg = generate_qr_svg(totp_uri)

            mfa_ticket = f"mfat_{uuid.uuid4().hex}"
            mfa_tickets[mfa_ticket] = {
                "user_id": user["id"],
                "username": user["username"],
                "mfa_secret": setup_secret,
                "is_setup": True,
                "expires_at": time.time() + 300,
            }
            return {
                "token": "",
                "mfa_required": True,
                "mfa_setup_required": True,
                "mfa_ticket": mfa_ticket,
                "qr_code": qr_svg,
                "secret": setup_secret,
                "must_change_password": bool(user["must_change_password"]),
                "user": {},
            }

        # Сброс неверных входов при успешной аутентификации
        conn.execute("UPDATE users SET last_login = CURRENT_TIMESTAMP, failed_login_attempts = 0, locked_until = NULL WHERE id = ?", (user["id"],))
        conn.commit()

        client_ip = request.client.host if request and request.client else "local"
        user_agent = request.headers.get("user-agent") if request else "Browser Session"
        token = create_access_token(user["id"], user["username"], client_ip, user_agent)
        must_change = bool(user["must_change_password"])

        perm_rows = conn.execute(
            "SELECT permission_id FROM role_permissions WHERE role_id = ?",
            (user["role_id"],),
        ).fetchall()
        perms = [p["permission_id"] for p in perm_rows]

        log_audit_event(
            user_id=user["id"],
            username=user["username"],
            action="auth.login_success",
            resource="auth",
            details=tr(request, "successful_login"),
            ip_address=request.client.host if request.client else None,
        )

        return {
            "token": token,
            "must_change_password": must_change,
            "mfa_required": False,
            "user": {
                "id": user["id"],
                "username": user["username"],
                "full_name": user["full_name"],
                "email": user["email"],
                "uid": user["uid"],
                "role_id": user["role_id"],
                "role_name": user["role_name"],
                "avatar": user["avatar"],
                "permissions": perms,
                "must_change_password": must_change,
            },
        }
    finally:
        conn.close()


@router.post("/auth/mfa/verify", response_model=LoginResponse)
async def verify_mfa_login(body: MfaVerifyRequest, request: Request):
    """Подтверждение шага MFA по мфа-билету и 6-значному коду."""
    ticket_info = mfa_tickets.get(body.mfa_ticket)
    if not ticket_info or time.time() > ticket_info["expires_at"]:
        raise AuthenticationError(message=tr(request, "login_session_expired"), code="LOGIN_SESSION_EXPIRED")

    user_id = ticket_info["user_id"]
    mfa_secret = ticket_info["mfa_secret"]

    conn = get_db_connection()
    try:
        user = conn.execute(
            """
            SELECT u.id, u.username, u.full_name, u.email, u.uid, u.avatar, u.role_id, r.name as role_name,
                   u.must_change_password, u.mfa_secret
            FROM users u
            JOIN roles r ON u.role_id = r.id
            WHERE u.id = ?
            """,
            (user_id,),
        ).fetchone()

        if not user:
            raise NotFoundError(message=tr(request, "user_not_found"), code="USER_NOT_FOUND")

        secret_to_check = mfa_secret or user["mfa_secret"]
        if not secret_to_check or not verify_totp_code(secret_to_check, body.code):
            log_audit_event(
                user_id=user["id"],
                username=user["username"],
                action="auth.mfa_failed",
                resource="auth",
                details=tr(request, "invalid_2fa_code"),
                ip_address=request.client.host if request.client else None,
            )
            raise AuthenticationError(message=tr(request, "invalid_2fa_code"), code="INVALID_2FA_CODE")

        is_setup = ticket_info.get("is_setup", False)
        mfa_tickets.pop(body.mfa_ticket, None)

        if is_setup:
            conn.execute(
                "UPDATE users SET mfa_enabled = 1, mfa_secret = ? WHERE id = ?",
                (secret_to_check, user["id"]),
            )
            log_audit_event(
                user_id=user["id"],
                username=user["username"],
                action="user.mfa_enabled",
                resource="user",
                details=tr(request, "2fa_forcibly_enabled"),
                ip_address=request.client.host if request and request.client else None,
            )

        conn.execute("UPDATE users SET last_login = CURRENT_TIMESTAMP, failed_login_attempts = 0, locked_until = NULL WHERE id = ?", (user["id"],))
        conn.commit()

        client_ip = request.client.host if request and request.client else "local"
        user_agent = request.headers.get("user-agent") if request else "Browser Session"
        token = create_access_token(user["id"], user["username"], client_ip, user_agent)
        must_change = bool(user["must_change_password"])

        perm_rows = conn.execute(
            "SELECT permission_id FROM role_permissions WHERE role_id = ?",
            (user["role_id"],),
        ).fetchall()
        perms = [p["permission_id"] for p in perm_rows]

        log_audit_event(
            user_id=user["id"],
            username=user["username"],
            action="auth.login_success",
            resource="auth",
            details=tr(request, "successful_login_mfa"),
            ip_address=request.client.host if request.client else None,
        )

        return {
            "token": token,
            "must_change_password": must_change,
            "mfa_required": False,
            "user": {
                "id": user["id"],
                "username": user["username"],
                "full_name": user["full_name"],
                "email": user["email"],
                "uid": user["uid"],
                "role_id": user["role_id"],
                "role_name": user["role_name"],
                "avatar": user["avatar"],
                "permissions": perms,
                "must_change_password": must_change,
            },
        }
    finally:
        conn.close()


@router.post("/auth/mfa/setup")
async def setup_mfa(current_user: CurrentUser = Depends(get_current_user)):
    """Генерация нового TOTP секрета и QR кода для текущего пользователя."""
    secret = generate_totp_secret()
    totp_uri = get_totp_uri(current_user.username, secret)
    qr_svg = generate_qr_svg(totp_uri)

    return {
        "secret": secret,
        "qr_code": qr_svg,
        "uri": totp_uri,
    }


@router.post("/auth/mfa/enable")
async def enable_mfa(
    body: MfaEnableRequest,
    current_user: CurrentUser = Depends(get_current_user),
    request: Request = None,
):
    """Подтверждение и активация 2FA в аккаунте."""
    if not verify_totp_code(body.secret, body.code):
        raise ValidationError(message=tr(request, "invalid_mfa_code"), code="INVALID_MFA_CODE")

    conn = get_db_connection()
    try:
        conn.execute(
            "UPDATE users SET mfa_enabled = 1, mfa_secret = ? WHERE id = ?",
            (body.secret, current_user.id),
        )
        conn.commit()

        log_audit_event(
            user_id=current_user.id,
            username=current_user.username,
            action="user.mfa_enabled",
            resource="user",
            details=tr(request, "2fa_enabled"),
            ip_address=request.client.host if request and request.client else None,
        )
        return {"ok": True}
    finally:
        conn.close()


@router.post("/auth/mfa/disable")
async def disable_mfa(
    current_user: CurrentUser = Depends(get_current_user),
    request: Request = None,
):
    """Отключение двухфакторной аутентификации."""
    sec_settings = get_security_settings()
    if bool(sec_settings.get("force_mfa", False)):
        raise ValidationError(message=tr(request, "disable_2fa_prohibited"), code="DISABLE_2FA_PROHIBITED")

    conn = get_db_connection()
    try:
        conn.execute(
            "UPDATE users SET mfa_enabled = 0, mfa_secret = NULL WHERE id = ?",
            (current_user.id,),
        )
        conn.commit()

        log_audit_event(
            user_id=current_user.id,
            username=current_user.username,
            action="user.mfa_disabled",
            resource="user",
            details=tr(request, "2fa_disabled"),
            ip_address=request.client.host if request and request.client else None,
        )
        return {"ok": True}
    finally:
        conn.close()


@router.post("/auth/logout")
async def logout(current_user: CurrentUser = Depends(get_current_user), request: Request = None):
    """Выход пользователя из системы."""
    conn = get_db_connection()
    try:
        auth_header = request.headers.get("authorization") if request else None
        if auth_header and auth_header.startswith("Bearer "):
            token = auth_header.split(" ", 1)[1]
            payload = decode_access_token(token)
            if payload and "jti" in payload:
                conn.execute("UPDATE active_sessions SET is_revoked = 1 WHERE token_jti = ?", (payload["jti"],))
                conn.commit()
    except Exception:
        pass
    finally:
        conn.close()

    log_audit_event(
        user_id=current_user.id,
        username=current_user.username,
        action="auth.logout",
        resource="auth",
        details=tr(request, "logged_out"),
        ip_address=request.client.host if request and request.client else None,
    )
    return {"ok": True}


@router.post("/auth/terminate-sessions")
async def terminate_all_sessions(
    request: Request = None,
    other_only: bool = False,
    current_user: CurrentUser = Depends(get_current_user),
):
    """Завершение всех активных сессий пользователя на всех устройствах."""
    conn = get_db_connection()
    try:
        import time
        now_ts = int(time.time())
        current_jti = current_user.token_jti
        if not current_jti and request:
            auth_header = request.headers.get("authorization")
            if auth_header and auth_header.startswith("Bearer "):
                payload = decode_access_token(auth_header.split(" ", 1)[1])
                if payload:
                    current_jti = payload.get("jti")

        if other_only and current_jti:
            conn.execute(
                "UPDATE active_sessions SET is_revoked = 1 WHERE user_id = ? AND token_jti != ?",
                (current_user.id, current_jti),
            )
        else:
            conn.execute(
                "UPDATE active_sessions SET is_revoked = 1 WHERE user_id = ?",
                (current_user.id,),
            )
            conn.execute(
                "UPDATE users SET token_valid_after = ? WHERE id = ?",
                (now_ts, current_user.id),
            )
        conn.commit()

        log_audit_event(
            user_id=current_user.id,
            username=current_user.username,
            action="auth.terminate_all_sessions",
            resource="auth",
            details=tr(request, "user_sessions_terminated"),
            ip_address=request.client.host if request and request.client else None,
        )
        return {"ok": True}
    finally:
        conn.close()


@router.get("/auth/me")
async def get_me(current_user: CurrentUser = Depends(get_current_user)):
    """Получение информации о текущем авторизованном пользователе."""
    sec_settings = get_security_settings()
    conn = get_db_connection()
    mfa_enabled = False
    try:
        u = conn.execute("SELECT mfa_enabled FROM users WHERE id = ?", (current_user.id,)).fetchone()
        if u and "mfa_enabled" in u.keys():
            mfa_enabled = bool(u["mfa_enabled"])
    finally:
        conn.close()

    return {
        "id": current_user.id,
        "username": current_user.username,
        "full_name": current_user.full_name,
        "email": current_user.email,
        "uid": current_user.uid,
        "role_id": current_user.role_id,
        "role_name": current_user.role_name,
        "avatar": current_user.avatar,
        "permissions": current_user.permissions,
        "auth_enabled": sec_settings.get("auth_enabled", True),
        "mfa_enabled": mfa_enabled,
        "force_mfa": bool(sec_settings.get("force_mfa", False)),
    }


# ── 2. Users Management API ──────────────────────────────────────────
@router.get("/users")
async def list_users(
    page: int = 1,
    page_size: int = 100,
    search: Optional[str] = None,
    role_id: Optional[str] = None,
    current_user: CurrentUser = Depends(require_permission("users.view")),
):
    """Получение списка всех пользователей с поддержкой пагинации, поиска и статуса активности."""
    conn = get_db_connection()
    try:
        sec_settings = get_security_settings()
        inactivity_timeout = int(sec_settings.get("inactivity_timeout_mins", 30))

        where_clauses = []
        params = []
        if search and search.strip():
            s = f"%{search.strip()}%"
            where_clauses.append("(u.username LIKE ? OR u.full_name LIKE ? OR u.email LIKE ? OR u.title LIKE ? OR u.uid LIKE ?)")
            params.extend([s, s, s, s, s])
        if role_id:
            where_clauses.append("u.role_id = ?")
            params.append(role_id)

        where_str = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

        total = conn.execute(f"SELECT COUNT(*) as cnt FROM users u {where_str}", params).fetchone()["cnt"]

        offset = max(0, (page - 1) * page_size)
        query_params = list(params) + [page_size, offset]

        rows = conn.execute(
            f"""
            SELECT u.id, u.username, u.full_name, u.email, u.title, u.uid, u.avatar, u.is_active, u.role_id, r.name as role_name,
                   u.created_at, u.last_login, u.last_seen, u.must_change_password, u.failed_login_attempts, u.locked_until
            FROM users u
            JOIN roles r ON u.role_id = r.id
            {where_str}
            ORDER BY u.created_at DESC
            LIMIT ? OFFSET ?
            """,
            query_params,
        ).fetchall()

        now = datetime.datetime.now(datetime.timezone.utc)
        items = []
        for r in rows:
            u_dict = dict(r)
            # Динамическое вычисление реального онлайн-статуса:
            # Пользователь "В сети" только если аккаунт активен И есть неаннулированная сессия с активостью за последние N минут
            is_online = False
            if u_dict.get("is_active"):
                inactivity_seconds = inactivity_timeout * 60
                active_sess = conn.execute(
                    """
                    SELECT COUNT(*) as cnt
                    FROM active_sessions
                    WHERE user_id = ? AND is_revoked = 0
                      AND (julianday('now') - julianday(replace(last_seen, 'T', ' '))) * 86400 <= ?
                    """,
                    (u_dict["id"], inactivity_seconds),
                ).fetchone()
                if active_sess and active_sess["cnt"] > 0:
                    is_online = True

            u_dict["is_online"] = is_online
            items.append(u_dict)

        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "items": items,
        }
    finally:
        conn.close()


@router.post("/users")
async def create_user(
    body: UserCreateRequest,
    current_user: CurrentUser = Depends(require_permission("users.manage")),
    request: Request = None,
):
    """Создание нового пользователя."""
    conn = get_db_connection()
    try:
        user_uid = body.uid.strip() if (body.uid and body.uid.strip()) else f"UID-{uuid.uuid4().hex[:6].upper()}"

        # Проверка сложности пароля
        validate_password_complexity(body.password, request)

        # Проверка уникальности username и uid
        existing = conn.execute(
            "SELECT username, uid FROM users WHERE username = ? OR uid = ?",
            (body.username, user_uid),
        ).fetchone()
        if existing:
            raise ValidationError(message=tr(request, "user_already_exists"), code="USER_ALREADY_EXISTS")

        new_id = f"usr-{uuid.uuid4().hex[:8]}"
        hashed_pass = hash_password(body.password)

        sec_settings = get_security_settings()
        must_change = body.must_change_password if body.must_change_password is not None else bool(sec_settings.get("mandatory_password_change", False))

        conn.execute(
            """
            INSERT INTO users (id, username, full_name, email, title, uid, hashed_password, is_active, role_id, must_change_password)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (new_id, body.username, body.full_name, body.email, body.title or "", user_uid, hashed_pass, int(body.is_active), body.role_id, int(must_change)),
        )
        conn.commit()

        log_audit_event(
            user_id=current_user.id,
            username=current_user.username,
            action="user.create",
            resource=f"user:{new_id}",
            details=tr(request, "created_user_audit", username=body.username, full_name=body.full_name),
            ip_address=request.client.host if request and request.client else None,
        )
        return {"ok": True, "id": new_id}
    finally:
        conn.close()


@router.put("/users/me")
async def update_own_profile(
    body: SelfUpdateRequest,
    current_user: CurrentUser = Depends(get_current_user),
    request: Request = None,
):
    """Обновление собственного профиля текущим пользователем."""
    conn = get_db_connection()
    try:
        updates = []
        params = []
        if body.full_name is not None:
            updates.append("full_name = ?")
            params.append(body.full_name)
        if body.email is not None:
            updates.append("email = ?")
            params.append(body.email)
        if body.avatar is not None:
            updates.append("avatar = ?")
            params.append(body.avatar)

        if updates:
            params.append(current_user.id)
            conn.execute(f"UPDATE users SET {', '.join(updates)} WHERE id = ?", params)
            conn.commit()

        log_audit_event(
            user_id=current_user.id,
            username=current_user.username,
            action="user.update_profile",
            resource="profile",
            details=tr(request, "updated_user_profile_audit"),
            ip_address=request.client.host if request and request.client else None,
        )
        return {"ok": True}
    finally:
        conn.close()


@router.put("/users/me/password")
async def change_own_password(
    body: PasswordChangeRequest,
    current_user: CurrentUser = Depends(get_current_user),
    request: Request = None,
):
    """Смена собственного пароля."""
    conn = get_db_connection()
    try:
        user = conn.execute("SELECT hashed_password FROM users WHERE id = ?", (current_user.id,)).fetchone()
        if not verify_password(body.old_password, user["hashed_password"]):
            raise ValidationError(message=tr(request, "current_password_incorrect"), code="CURRENT_PASSWORD_INCORRECT")

        validate_password_complexity(body.new_password, request)

        new_hash = hash_password(body.new_password)
        conn.execute("UPDATE users SET hashed_password = ?, must_change_password = 0 WHERE id = ?", (new_hash, current_user.id))
        conn.commit()

        log_audit_event(
            user_id=current_user.id,
            username=current_user.username,
            action="user.change_password",
            resource="profile",
            details=tr(request, "user_changed_password_audit"),
            ip_address=request.client.host if request and request.client else None,
        )
        return {"ok": True}
    finally:
        conn.close()


@router.put("/users/{user_id}")
async def update_user(
    user_id: str,
    body: UserUpdateRequest,
    current_user: CurrentUser = Depends(require_permission("users.manage")),
    request: Request = None,
):
    """Редактирование пользователя."""
    conn = get_db_connection()
    try:
        user = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if not user:
            raise NotFoundError(message=tr(request, "user_not_found"), code="USER_NOT_FOUND")

        updates = []
        params = []
        if body.full_name is not None:
            updates.append("full_name = ?")
            params.append(body.full_name)
        if body.email is not None:
            updates.append("email = ?")
            params.append(body.email)
        if body.title is not None:
            updates.append("title = ?")
            params.append(body.title)
        if body.role_id is not None:
            updates.append("role_id = ?")
            params.append(body.role_id)
        if body.is_active is not None:
            updates.append("is_active = ?")
            params.append(int(body.is_active))
            if body.is_active:
                updates.append("failed_login_attempts = 0")
                updates.append("locked_until = NULL")
        if body.must_change_password is not None:
            updates.append("must_change_password = ?")
            params.append(int(body.must_change_password))
        if body.password and body.password.strip():
            validate_password_complexity(body.password, request)
            updates.append("hashed_password = ?")
            params.append(hash_password(body.password))

        # Проверка защиты root от отключения / смены роли
        if (body.is_active is False or (body.role_id and body.role_id != '1')) and (user["username"] == "root" or user["role_id"] == "1"):
            other_superusers = conn.execute(
                "SELECT COUNT(*) as cnt FROM users WHERE role_id = '1' AND is_active = 1 AND id != ?",
                (user_id,),
            ).fetchone()["cnt"]
            if other_superusers == 0:
                raise ValidationError(message=tr(request, "cannot_disable_root"), code="CANNOT_DISABLE_ROOT")

        if updates:
            params.append(user_id)
            conn.execute(f"UPDATE users SET {', '.join(updates)} WHERE id = ?", params)
            conn.commit()

        log_audit_event(
            user_id=current_user.id,
            username=current_user.username,
            action="user.update",
            resource=f"user:{user_id}",
            details=tr(request, "updated_user_audit", username=user['username']),
            ip_address=request.client.host if request and request.client else None,
        )
        return {"ok": True}
    finally:
        conn.close()


@router.delete("/users/{user_id}")
async def delete_user(
    user_id: str,
    current_user: CurrentUser = Depends(require_permission("users.manage")),
    request: Request = None,
):
    """Удаление пользователя."""
    if user_id == current_user.id:
        raise ValidationError(message=tr(request, "cannot_delete_self"), code="CANNOT_DELETE_SELF")

    conn = get_db_connection()
    try:
        user = conn.execute("SELECT username, role_id FROM users WHERE id = ?", (user_id,)).fetchone()
        if not user:
            raise NotFoundError(message=tr(request, "user_not_found"), code="USER_NOT_FOUND")

        if user["username"] == "root" or user["role_id"] == "1":
            other_superusers = conn.execute(
                "SELECT COUNT(*) as cnt FROM users WHERE role_id = '1' AND is_active = 1 AND id != ?",
                (user_id,),
            ).fetchone()["cnt"]
            if other_superusers == 0:
                raise ValidationError(message=tr(request, "cannot_delete_root"), code="CANNOT_DELETE_ROOT")

        conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
        conn.commit()

        log_audit_event(
            user_id=current_user.id,
            username=current_user.username,
            action="user.delete",
            resource=f"user:{user_id}",
            details=tr(request, "deleted_user_audit", username=user['username']),
            ip_address=request.client.host if request and request.client else None,
        )
        return {"ok": True}
    finally:
        conn.close()


@router.post("/users/{user_id}/terminate-sessions")
async def terminate_user_sessions(
    user_id: str,
    current_user: CurrentUser = Depends(require_permission("users.manage")),
    request: Request = None,
):
    """Принудительное завершение всех сессий выбранного пользователя администратором."""
    conn = get_db_connection()
    try:
        user = conn.execute("SELECT username FROM users WHERE id = ?", (user_id,)).fetchone()
        if not user:
            raise NotFoundError(message=tr(request, "user_not_found"), code="USER_NOT_FOUND")

        import time
        now_ts = int(time.time())
        conn.execute("UPDATE active_sessions SET is_revoked = 1 WHERE user_id = ?", (user_id,))
        conn.execute("UPDATE users SET token_valid_after = ? WHERE id = ?", (now_ts, user_id))
        conn.commit()

        log_audit_event(
            user_id=current_user.id,
            username=current_user.username,
            action="user.terminate_sessions",
            resource=f"user:{user_id}",
            details=tr(request, "admin_terminated_user_sessions_audit", username=user['username']),
            ip_address=request.client.host if request and request.client else None,
        )
        return {"ok": True}
    finally:
        conn.close()


# ── 3. Roles & Permissions API (RBAC) ────────────────────────────────
@router.get("/roles")
async def list_roles(current_user: CurrentUser = Depends(require_permission("roles.view"))):
    """Получение списка ролей с привязанными правами."""
    conn = get_db_connection()
    try:
        roles_rows = conn.execute(
            """
            SELECT r.id, r.name, r.description, r.is_system, COUNT(u.id) as users_count
            FROM roles r
            LEFT JOIN users u ON u.role_id = r.id
            GROUP BY r.id
            """
        ).fetchall()

        res = []
        for r in roles_rows:
            perm_rows = conn.execute(
                "SELECT permission_id FROM role_permissions WHERE role_id = ?",
                (r["id"],),
            ).fetchall()
            perms = [p["permission_id"] for p in perm_rows]
            res.append({
                "id": r["id"],
                "name": r["name"],
                "description": r["description"],
                "is_system": bool(r["is_system"]),
                "users_count": r["users_count"],
                "permissions": perms,
            })
        return res
    finally:
        conn.close()


@router.get("/permissions")
async def list_permissions(current_user: CurrentUser = Depends(require_permission("roles.view"))):
    """Получение списка всех доступных системных и модульных разрешений."""
    conn = get_db_connection()
    try:
        rows = conn.execute("SELECT id, category, name, description, module_id FROM permissions ORDER BY category ASC, id ASC").fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@router.post("/roles")
async def create_role(
    body: RoleCreateUpdateRequest,
    current_user: CurrentUser = Depends(require_permission("roles.manage")),
    request: Request = None,
):
    """Создание новой роли."""
    conn = get_db_connection()
    try:
        new_id = str(uuid.uuid4().hex[:8])
        conn.execute(
            "INSERT INTO roles (id, name, description, is_system) VALUES (?, ?, ?, 0)",
            (new_id, body.name, body.description),
        )
        for pid in body.permission_ids:
            conn.execute(
                "INSERT INTO role_permissions (role_id, permission_id) VALUES (?, ?)",
                (new_id, pid),
            )
        conn.commit()

        log_audit_event(
            user_id=current_user.id,
            username=current_user.username,
            action="role.create",
            resource=f"role:{new_id}",
            details=tr(request, "created_role_audit", name=body.name),
            ip_address=request.client.host if request and request.client else None,
        )
        return {"ok": True, "id": new_id}
    finally:
        conn.close()


@router.put("/roles/{role_id}")
async def update_role(
    role_id: str,
    body: RoleCreateUpdateRequest,
    current_user: CurrentUser = Depends(require_permission("roles.manage")),
    request: Request = None,
):
    """Обновление роли и матрицы ее разрешений."""
    conn = get_db_connection()
    try:
        role = conn.execute("SELECT * FROM roles WHERE id = ?", (role_id,)).fetchone()
        if not role:
            raise NotFoundError(message=tr(request, "role_not_found"), code="ROLE_NOT_FOUND")

        conn.execute(
            "UPDATE roles SET name = ?, description = ? WHERE id = ?",
            (body.name, body.description, role_id),
        )
        # Перезапись разрешений
        conn.execute("DELETE FROM role_permissions WHERE role_id = ?", (role_id,))
        for pid in body.permission_ids:
            conn.execute(
                "INSERT INTO role_permissions (role_id, permission_id) VALUES (?, ?)",
                (role_id, pid),
            )
        conn.commit()
        clear_permissions_cache(role_id)

        log_audit_event(
            user_id=current_user.id,
            username=current_user.username,
            action="role.update",
            resource=f"role:{role_id}",
            details=tr(request, "updated_role_audit", name=body.name),
            ip_address=request.client.host if request and request.client else None,
        )
        return {"ok": True}
    finally:
        conn.close()


@router.delete("/roles/{role_id}")
async def delete_role(
    role_id: str,
    current_user: CurrentUser = Depends(require_permission("roles.manage")),
    request: Request = None,
):
    """Удаление пользовательской роли."""
    conn = get_db_connection()
    try:
        role = conn.execute("SELECT name, is_system FROM roles WHERE id = ?", (role_id,)).fetchone()
        if not role:
            raise NotFoundError(message=tr(request, "role_not_found"), code="ROLE_NOT_FOUND")

        if role["is_system"]:
            raise ValidationError(message=tr(request, "cannot_delete_system_role"), code="CANNOT_DELETE_SYSTEM_ROLE")

        assigned_users = conn.execute("SELECT COUNT(*) as cnt FROM users WHERE role_id = ?", (role_id,)).fetchone()["cnt"]
        if assigned_users > 0:
            raise ValidationError(message=tr(request, "cannot_delete_assigned_role", name=role['name'], assigned_users=assigned_users), code="ROLE_IN_USE")

        conn.execute("DELETE FROM role_permissions WHERE role_id = ?", (role_id,))
        conn.execute("DELETE FROM roles WHERE id = ?", (role_id,))
        conn.commit()
        clear_permissions_cache(role_id)

        log_audit_event(
            user_id=current_user.id,
            username=current_user.username,
            action="role.delete",
            resource=f"role:{role_id}",
            details=tr(request, "deleted_role_audit", name=role['name']),
            ip_address=request.client.host if request and request.client else None,
        )
        return {"ok": True}
    finally:
        conn.close()


# ── 4. Audit Logs API ────────────────────────────────────────────────
@router.get("/audit-logs")
async def get_audit_logs(
    limit: int = 100,
    offset: int = 0,
    category: Optional[str] = None,
    search: Optional[str] = None,
    current_user: CurrentUser = Depends(require_permission("audit.view")),
):
    """Получение журнала событий аудита с поддержкой серверной фильтрации."""
    conn = get_db_connection()
    try:
        where_clauses = []
        params = []
        if category == "errors":
            where_clauses.append("(action LIKE '%failed%' OR action LIKE '%delete%' OR action LIKE '%lockout%')")
        elif category == "auth":
            where_clauses.append("action LIKE 'auth.%'")
        elif category == "user":
            where_clauses.append("(action LIKE 'user.%' OR action LIKE 'role.%')")

        if search and search.strip():
            s = f"%{search.strip()}%"
            where_clauses.append("(username LIKE ? OR action LIKE ? OR resource LIKE ? OR details LIKE ? OR ip_address LIKE ?)")
            params.extend([s, s, s, s, s])

        where_str = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

        total = conn.execute(f"SELECT COUNT(*) as cnt FROM audit_logs {where_str}", params).fetchone()["cnt"]

        query_params = list(params) + [limit, offset]
        rows = conn.execute(
            f"""
            SELECT id, timestamp, user_id, username, action, resource, details, ip_address
            FROM audit_logs
            {where_str}
            ORDER BY id DESC
            LIMIT ? OFFSET ?
            """,
            query_params,
        ).fetchall()

        return {
            "total": total,
            "items": [dict(r) for r in rows],
        }
    finally:
        conn.close()


def generate_audit_excel(rows) -> bytes:
    """Генерация стилизованного отформатированного файла Excel (.xlsx) журнала аудита."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Журнал аудита"
    ws.views.sheetView[0].showGridLines = True

    headers = ["ID", "Дата и время", "Пользователь", "Действие", "Ресурс", "Детали события", "IP-адрес"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    row_font = Font(name="Calibri", size=10, color="0F172A")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    ws.row_dimensions[1].height = 28
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    for idx, r in enumerate(rows, start=2):
        ws.append([
            r["id"],
            r["timestamp"],
            r["username"],
            r["action"],
            r["resource"],
            r["details"] or "",
            r["ip_address"] or "",
        ])
        ws.row_dimensions[idx].height = 20
        use_alt = (idx % 2 == 0)

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=idx, column=col_num)
            cell.font = row_font
            cell.border = thin_border
            if use_alt:
                cell.fill = row_alt_fill

            if col_num in (1, 2, 7):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A2"

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/audit-logs/export")
async def export_audit_logs(
    format: str = "xlsx",
    current_user: CurrentUser = Depends(require_permission("audit.export")),
):
    """Экспорт журнала событий аудита в формат Excel (.xlsx) с форматированием или CSV."""
    conn = get_db_connection()
    try:
        rows = conn.execute(
            """
            SELECT id, timestamp, username, action, resource, details, ip_address
            FROM audit_logs
            ORDER BY id DESC
            """
        ).fetchall()

        if format == "csv":
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(["ID", "Timestamp", "Username", "Action", "Resource", "Details", "IP Address"])
            for r in rows:
                writer.writerow([r["id"], r["timestamp"], r["username"], r["action"], r["resource"], r["details"] or "", r["ip_address"] or ""])

            return Response(
                content=output.getvalue().encode("utf-8-sig"),
                media_type="text/csv",
                headers={"Content-Disposition": 'attachment; filename="audit_logs.csv"'},
            )

        xlsx_bytes = generate_audit_excel(rows)
        return Response(
            content=xlsx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="audit_logs.xlsx"'},
        )
    finally:
        conn.close()


class AuditRotateRequest(BaseModel):
    max_days: int = 90
    max_records: int = 100000


@router.post("/audit-logs/rotate")
async def rotate_audit_logs_endpoint(
    body: Optional[AuditRotateRequest] = None,
    request: Request = None,
    current_user: CurrentUser = Depends(require_permission("system.admin")),
):
    """Принудительная очистка/ротация устаревших логов аудита."""
    from backend.core.audit import rotate_audit_logs
    max_days = body.max_days if body else 90
    max_records = body.max_records if body else 100000

    deleted = rotate_audit_logs(max_days=max_days, max_records=max_records)
    log_audit_event(
        user_id=current_user.id,
        username=current_user.username,
        action="system.audit_logs_rotated",
        resource="audit",
        details=tr(request, "audit_logs_rotated", deleted=deleted),
        ip_address=request.client.host if request and request.client else None,
    )
    return {"ok": True, "deleted_count": deleted}


def validate_password_complexity(password: str, request: Request = None) -> None:
    """Проверка пароля на соответствие системной политике сложности."""
    if not password:
        return
    sec_settings = get_security_settings()
    min_len = int(sec_settings.get("min_password_length", 8))
    req_upper = bool(sec_settings.get("require_uppercase", False))
    req_digits = bool(sec_settings.get("require_digits", False))
    req_special = bool(sec_settings.get("require_special_chars", False))

    if len(password) < min_len:
        raise ValidationError(message=tr(request, "password_too_short", min_len=min_len), code="PASSWORD_TOO_SHORT", details={"min_len": min_len})
    if req_upper and not any(c.isupper() for c in password):
        raise ValidationError(message=tr(request, "password_require_uppercase"), code="PASSWORD_REQUIRE_UPPERCASE")
    if req_digits and not any(c.isdigit() for c in password):
        raise ValidationError(message=tr(request, "password_require_digits"), code="PASSWORD_REQUIRE_DIGITS")
    if req_special and not any(c in "!@#$%^&*()_+-=[]{}|;:,.<>?" for c in password):
        raise ValidationError(message=tr(request, "password_require_special"), code="PASSWORD_REQUIRE_SPECIAL")


# ── 5. System Security Settings API ────────────────────────────────
class SecuritySettingsModel(BaseModel):
    auth_enabled: bool = True
    mandatory_password_change: bool = True
    max_login_attempts: int = 5
    lockout_duration: int = 30
    session_ttl_hours: int = 12
    inactivity_timeout_mins: int = 30
    force_mfa: bool = False
    min_password_length: int = 8
    require_uppercase: bool = False
    require_digits: bool = False
    require_special_chars: bool = False
    ip_whitelist: Optional[str] = ""


@router.get("/settings/security", response_model=SecuritySettingsModel)
async def get_security_settings_endpoint(
    current_user: CurrentUser = Depends(require_permission("settings.view")),
):
    """Получение глобальных настроек безопасности."""
    return get_security_settings()


@router.put("/settings/security")
async def update_security_settings_endpoint(
    body: SecuritySettingsModel,
    request: Request,
    current_user: CurrentUser = Depends(require_permission("settings.edit")),
):
    """Обновление глобальных настроек безопасности."""
    save_security_settings(body.model_dump())
    log_audit_event(
        user_id=current_user.id,
        username=current_user.username,
        action="system.security_settings_updated",
        resource="settings",
        details=tr(request, "Обновлены параметры политики безопасности", "Updated security policy parameters"),
        ip_address=request.client.host if request and request.client else None,
    )
    return {"ok": True}


# ── 6. Active Sessions & Bulk Users API ───────────────────────────
class BulkUsersActionRequest(BaseModel):
    user_ids: List[str]
    action: str  # lock, unlock, set_role, terminate_sessions
    role_id: Optional[str] = None


@router.get("/users/me/sessions")
async def get_my_sessions(
    request: Request = None,
    current_user: CurrentUser = Depends(get_current_user),
):
    """Получение списка собственных активных сессий текущего пользователя."""
    sec_settings = get_security_settings()
    ttl_hours = int(sec_settings.get("session_ttl_hours", 12))
    ttl_seconds = ttl_hours * 3600
    current_jti = None
    if request:
        auth_header = request.headers.get("authorization")
        if auth_header and auth_header.startswith("Bearer "):
            payload = decode_access_token(auth_header.split(" ", 1)[1])
            if payload:
                current_jti = payload.get("jti")

    conn = get_db_connection()
    try:
        rows = conn.execute(
            """
            SELECT id, token_jti, ip_address, user_agent, created_at, last_seen, is_revoked
            FROM active_sessions
            WHERE user_id = ? AND is_revoked = 0
              AND (julianday('now') - julianday(replace(last_seen, 'T', ' '))) * 86400 <= ?
            ORDER BY last_seen DESC
            """,
            (current_user.id, ttl_seconds),
        ).fetchall()

        result = []
        for r in rows:
            d = dict(r)
            d["is_current"] = bool(current_jti and d.get("token_jti") == current_jti)
            result.append(d)
        return result
    finally:
        conn.close()


@router.delete("/users/me/sessions/{session_id}")
async def revoke_my_session(
    session_id: str,
    current_user: CurrentUser = Depends(get_current_user),
    request: Request = None,
):
    """Аннулирование собственной сессии пользователя."""
    conn = get_db_connection()
    try:
        conn.execute("UPDATE active_sessions SET is_revoked = 1 WHERE id = ? AND user_id = ?", (session_id, current_user.id))
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()


@router.get("/users/{user_id}/sessions")
async def get_user_sessions(
    user_id: str,
    current_user: CurrentUser = Depends(require_permission("users.manage")),
    request: Request = None,
):
    """Получение списка активных сессий конкретного пользователя."""
    sec_settings = get_security_settings()
    ttl_hours = int(sec_settings.get("session_ttl_hours", 12))
    ttl_seconds = ttl_hours * 3600
    conn = get_db_connection()
    try:
        rows = conn.execute(
            """
            SELECT id, ip_address, user_agent, created_at, last_seen, is_revoked
            FROM active_sessions
            WHERE user_id = ? AND is_revoked = 0
              AND (julianday('now') - julianday(replace(last_seen, 'T', ' '))) * 86400 <= ?
            ORDER BY last_seen DESC
            """,
            (user_id, ttl_seconds),
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@router.delete("/users/sessions/{session_id}")
async def revoke_session(
    session_id: str,
    current_user: CurrentUser = Depends(require_permission("users.manage")),
    request: Request = None,
):
    """Точечное аннулирование выбранной сессии."""
    conn = get_db_connection()
    try:
        conn.execute("UPDATE active_sessions SET is_revoked = 1 WHERE id = ?", (session_id,))
        conn.commit()
        log_audit_event(
            user_id=current_user.id,
            username=current_user.username,
            action="auth.session_revoked",
            resource=f"session:{session_id}",
            details=tr(request, f"Аннулирована активная сессия {session_id}", f"Revoked active session {session_id}"),
            ip_address=request.client.host if request and request.client else None,
        )
        return {"ok": True}
    finally:
        conn.close()


@router.post("/users/bulk-action")
async def bulk_users_action(
    body: BulkUsersActionRequest,
    current_user: CurrentUser = Depends(require_permission("users.manage")),
    request: Request = None,
):
    """Массовые операции над выбранными пользователями."""
    if not body.user_ids:
        raise ValidationError(message=tr(request, "no_users_selected"), code="NO_USERS_SELECTED")

    conn = get_db_connection()
    try:
        if body.action == "lock":
            valid_ids = [uid for uid in body.user_ids if uid != "1"]
            if valid_ids:
                placeholders = ",".join(["?"] * len(valid_ids))
                conn.execute(f"UPDATE users SET is_active = 0 WHERE id IN ({placeholders}) AND username != 'root'", valid_ids)
                conn.commit()
        elif body.action == "unlock":
            placeholders = ",".join(["?"] * len(body.user_ids))
            conn.execute(f"UPDATE users SET is_active = 1, failed_login_attempts = 0, locked_until = NULL WHERE id IN ({placeholders})", body.user_ids)
            conn.commit()
        elif body.action == "set_role" and body.role_id:
            placeholders = ",".join(["?"] * len(body.user_ids))
            params = [body.role_id] + body.user_ids
            conn.execute(f"UPDATE users SET role_id = ? WHERE id IN ({placeholders}) AND username != 'root'", params)
            conn.commit()
        elif body.action == "terminate_sessions":
            import time
            now_ts = int(time.time())
            placeholders = ",".join(["?"] * len(body.user_ids))
            params = [now_ts] + body.user_ids
            conn.execute(f"UPDATE users SET token_valid_after = ? WHERE id IN ({placeholders})", params)
            conn.execute(f"UPDATE active_sessions SET is_revoked = 1 WHERE user_id IN ({placeholders})", body.user_ids)
            conn.commit()

        log_audit_event(
            user_id=current_user.id,
            username=current_user.username,
            action=f"user.bulk_{body.action}",
            resource="users",
            details=tr(request, "bulk_action", action=body.action, count=len(body.user_ids)),
            ip_address=request.client.host if request and request.client else None,
        )
        return {"ok": True, "count": len(body.user_ids)}
    finally:
        conn.close()

